from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nutrition_ingest.common.contracts import FieldKind, FieldSpec
from nutrition_ingest.common.units import (
    convert_numeric_value,
    extract_unit_from_label,
    sum_complete_numeric_components,
)
from nutrition_ingest.nutrient_mapping import AUSTRALIA_FIELD_SPECS, CORE_FIELD_UNITS

DEFAULT_SHEET_NAME = "All solids & liquids per 100 g"


def normalize_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        candidate = value.replace(",", "").strip()
        if not candidate:
            return None
        try:
            return float(candidate)
        except ValueError:
            return value.strip()
    return value


def canonicalize_header(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.replace("\n", " ").split()).casefold()


def specs_to_source_map(specs: dict[str, FieldSpec]) -> dict[str, str]:
    resolved: dict[str, str] = {}
    for field, spec in specs.items():
        if spec.kind in {FieldKind.SOURCE, FieldKind.LITERAL} and isinstance(spec.source, str):
            resolved[field] = spec.source
        else:
            resolved[field] = ""
    return resolved


def resolve_source_columns(headers: list[Any], mapping: dict[str, str]) -> dict[str, str | None]:
    header_values = [h for h in headers if isinstance(h, str)]
    headers_by_canonical = {canonicalize_header(h): h for h in header_values}

    resolved: dict[str, str | None] = {}
    for target_field, source_column in mapping.items():
        if source_column == "" or target_field in {"source", "portions"}:
            resolved[target_field] = source_column
            continue

        if source_column in header_values:
            resolved[target_field] = source_column
            continue

        source_canonical = canonicalize_header(source_column)
        direct_match = headers_by_canonical.get(source_canonical)
        if direct_match is not None:
            resolved[target_field] = direct_match
            continue

        startswith_match = next(
            (
                header
                for header in header_values
                if canonicalize_header(header).startswith(source_canonical)
            ),
            None,
        )
        resolved[target_field] = startswith_match

    return resolved


def build_source_units(resolved_mapping: dict[str, str | None]) -> dict[str, str | None]:
    units: dict[str, str | None] = {}
    for target_field, source_column in resolved_mapping.items():
        if target_field in {"source", "portions"}:
            continue
        if source_column in ("", None):
            continue
        units[target_field] = extract_unit_from_label(source_column)
    return units


def map_australia_row(
    raw_row: dict[str, Any],
    resolved_mapping: dict[str, str | None],
    source_units: dict[str, str | None],
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    for target_field, source_column in resolved_mapping.items():
        if target_field == "source":
            normalized[target_field] = source_column
            continue

        if target_field == "portions":
            normalized[target_field] = None
            continue

        if source_column in ("", None):
            normalized[target_field] = None
            continue

        value = normalize_value(raw_row.get(source_column))
        if isinstance(value, (int, float)):
            value = convert_numeric_value(
                value,
                source_units.get(target_field),
                CORE_FIELD_UNITS.get(target_field),
                field=target_field,
            )
        normalized[target_field] = value

    omega_components = [
        normalized.get(field) for field in ("omega_3_ala", "omega_3_epa", "omega_3_dha")
    ]
    normalized["omega_3_ala_epa_dha_sum"] = sum_complete_numeric_components(omega_components)

    return normalized


def iter_rows(
    workbook_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    field_specs: dict[str, FieldSpec] = AUSTRALIA_FIELD_SPECS,
):
    source_map = specs_to_source_map(field_specs)

    wb = load_workbook(filename=workbook_path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
    resolved_mapping = resolve_source_columns(headers, source_map)
    source_units = build_source_units(resolved_mapping)

    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_row = dict(zip(headers, row))
        mapped_row = map_australia_row(raw_row, resolved_mapping, source_units)
        if mapped_row.get("source_id") is None and mapped_row.get("name") is None:
            continue
        yield mapped_row
