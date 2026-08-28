from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nutrition_ingest.common.contracts import FieldKind, FieldSpec
from nutrition_ingest.common.units import (
    convert_numeric_value,
    extract_unit_from_label,
    sum_complete_numeric_components,
)
from nutrition_ingest.nutrient_mapping import CORE_FIELD_UNITS, NEVO_FIELD_SPECS

DEFAULT_SHEET_NAME = "NEVO2025"
DEFAULT_NUTRIENTS_SHEET_NAME = "NEVO2025_Nutrienten_Nutrients"
NEVO_COLUMN_POLYOLS = "POLYL"
NEVO_CODE_PREFIX_PATTERN = re.compile(r"^([A-Z0-9:_-]+)\s*(?:\(|$)")


def normalize_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        candidate = value.strip().replace(",", ".")
        if not candidate:
            return None
        try:
            return float(candidate)
        except ValueError:
            return value.strip()
    return value


def canonicalize_header(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.replace("\n", " ").strip().split()).casefold()


def extract_nevo_code(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    match = NEVO_CODE_PREFIX_PATTERN.match(value.strip())
    if match is None:
        return None
    return match.group(1)


def specs_to_source_map(specs: dict[str, FieldSpec]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field, spec in specs.items():
        if spec.kind in {FieldKind.SOURCE, FieldKind.LITERAL} and isinstance(spec.source, str):
            mapping[field] = spec.source
        else:
            mapping[field] = ""
    return mapping


def resolve_source_columns(headers: list[Any], mapping: dict[str, str]) -> dict[str, str | None]:
    header_values = [h for h in headers if isinstance(h, str)]
    headers_by_canonical = {canonicalize_header(h): h for h in header_values}
    headers_by_code = {
        code: header for header in header_values if (code := extract_nevo_code(header)) is not None
    }

    resolved: dict[str, str | None] = {}
    for target_field, source_column in mapping.items():
        if source_column == "" or target_field in {"source", "portions"}:
            resolved[target_field] = source_column
            continue

        if source_column in header_values:
            resolved[target_field] = source_column
            continue

        source_canonical = canonicalize_header(source_column)
        direct_match = headers_by_canonical.get(source_canonical)
        if direct_match is not None:
            resolved[target_field] = direct_match
            continue

        code_match = headers_by_code.get(source_column)
        if code_match is not None:
            resolved[target_field] = code_match
            continue

        startswith_match = next(
            (
                header
                for header in header_values
                if canonicalize_header(header).startswith(source_canonical)
            ),
            None,
        )
        resolved[target_field] = startswith_match

    return resolved


def get_nutrient_metadata(
    workbook_path: Path, nutrients_sheet_name: str
) -> dict[str, dict[str, str | None]]:
    wb = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    ws = wb[nutrients_sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    col_index = {canonicalize_header(h): i for i, h in enumerate(headers)}
    code_idx = col_index.get(canonicalize_header("Nutrient-code"))
    component_idx = col_index.get(canonicalize_header("Component"))
    unit_idx = col_index.get(canonicalize_header("Eenheid/Unit"))
    if code_idx is None:
        raise RuntimeError("Could not find Nutrient-code column in nutrients sheet")

    metadata: dict[str, dict[str, str | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code_raw = row[code_idx]
        if not isinstance(code_raw, str):
            continue
        code = code_raw.strip()
        if not code:
            continue
        component = (
            row[component_idx].strip()
            if component_idx is not None and isinstance(row[component_idx], str)
            else None
        )
        unit = (
            row[unit_idx].strip()
            if unit_idx is not None and isinstance(row[unit_idx], str)
            else None
        )
        metadata[code] = {"component": component, "unit": unit}

    return metadata


def validate_mapping_codes(
    mapping: dict[str, str], nutrient_metadata: dict[str, dict[str, str | None]]
) -> list[str]:
    skip_fields = {"source_id", "source", "name", "portions"}
    missing_codes: list[str] = []
    for target_field, source_column in mapping.items():
        if target_field in skip_fields or source_column == "":
            continue
        # Nutrient code columns in NEVO are short identifiers like CHO, VITC, F18:3CN3.
        if source_column in nutrient_metadata:
            continue
        if (
            target_field in {"carbohydrates_total", "carbohydrates_net_calculated"}
            and source_column == "CHO"
        ):
            continue
        if source_column in {"NEVO-code", "Engelse naam/Food name"}:
            continue
        missing_codes.append(source_column)
    return sorted(set(missing_codes))


def build_source_units_by_field(
    mapping: dict[str, str], nutrient_metadata: dict[str, dict[str, str | None]]
) -> dict[str, str | None]:
    source_units: dict[str, str | None] = {}
    for target_field, source_column in mapping.items():
        if target_field in {"source", "portions"} or source_column == "":
            continue
        if source_column in nutrient_metadata:
            source_units[target_field] = nutrient_metadata[source_column].get("unit")
            continue
        source_units[target_field] = extract_unit_from_label(source_column)
    return source_units


def compute_net_carbs(carbohydrates: Any, polyols: Any) -> float | None:
    if not isinstance(carbohydrates, (int, float)):
        return None
    if not isinstance(polyols, (int, float)):
        return None
    return max(0.0, float(carbohydrates) - float(polyols))


def map_nevo_row(
    raw_row: dict[str, Any],
    resolved_mapping: dict[str, str | None],
    source_units_by_field: dict[str, str | None],
    polyols_column: str | None,
    polyols_unit: str | None,
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    for target_field, source_column in resolved_mapping.items():
        if target_field == "source":
            normalized[target_field] = source_column
            continue
        if target_field == "portions":
            normalized[target_field] = None
            continue
        if source_column in ("", None):
            normalized[target_field] = None
            continue

        value = normalize_value(raw_row.get(source_column))
        if target_field == "source_id" and value is not None:
            # Keep source IDs stable as strings.
            if isinstance(value, (int, float)):
                normalized[target_field] = str(int(value))
            else:
                normalized[target_field] = str(value)
            continue

        if isinstance(value, (int, float)):
            value = convert_numeric_value(
                value,
                source_units_by_field.get(target_field),
                CORE_FIELD_UNITS.get(target_field),
                field=target_field,
            )
        normalized[target_field] = value

    polyols_value = normalize_value(raw_row.get(polyols_column)) if polyols_column else None
    if isinstance(polyols_value, (int, float)):
        polyols_value = convert_numeric_value(
            polyols_value,
            polyols_unit,
            CORE_FIELD_UNITS.get("carbohydrates_total"),
            field="carbohydrates_total",
        )
    normalized["carbohydrates_net_calculated"] = compute_net_carbs(
        normalized.get("carbohydrates_total"), polyols_value
    )
    omega_components = [
        normalized.get(field) for field in ("omega_3_ala", "omega_3_epa", "omega_3_dha")
    ]
    normalized["omega_3_ala_epa_dha_sum"] = sum_complete_numeric_components(omega_components)

    return normalized


def iter_rows(
    workbook_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    nutrients_sheet_name: str = DEFAULT_NUTRIENTS_SHEET_NAME,
    field_specs: dict[str, FieldSpec] = NEVO_FIELD_SPECS,
):
    mapping = specs_to_source_map(field_specs)
    nutrient_metadata = get_nutrient_metadata(workbook_path, nutrients_sheet_name)
    missing_codes = validate_mapping_codes(mapping, nutrient_metadata)
    if missing_codes:
        raise RuntimeError(
            "Mapping references nutrient codes not present in nutrients sheet: "
            + ", ".join(missing_codes)
        )

    wb = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    resolved_mapping = resolve_source_columns(headers, mapping)
    source_units_by_field = build_source_units_by_field(mapping, nutrient_metadata)
    polyols_column = resolve_source_columns(headers, {"polyols": NEVO_COLUMN_POLYOLS}).get(
        "polyols"
    )
    polyols_unit = nutrient_metadata.get(NEVO_COLUMN_POLYOLS, {}).get("unit")

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_row = dict(zip(headers, row))
        mapped_row = map_nevo_row(
            raw_row,
            resolved_mapping,
            source_units_by_field,
            polyols_column,
            polyols_unit,
        )
        if mapped_row.get("source_id") is None and mapped_row.get("name") is None:
            continue
        yield mapped_row
